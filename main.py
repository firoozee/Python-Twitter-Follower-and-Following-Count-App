import tweepy
import openpyxl
import schedule
import datetime
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Enter your Twitter API credentials here
consumer_key = "your_consumer_key"
consumer_secret = "your_consumer_secret"
access_token = "your_access_token"
access_token_secret = "your_access_token_secret"

# Enter your email credentials here
sender_email = "put your sender email here"
sender_password = "put your sender email app password here"
receiver_email = "put your reciver email here"

# Authenticate with Twitter API
auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_token_secret)
api = tweepy.API(auth)

# Enter the Twitter handles you want to get the followers and following count for here without @
twitter_handles = ["example_handle2", "example_handle2", "example_handle2"]

def get_follower_counts():
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set the header row
    worksheet.cell(row=1, column=1, value="Twitter Handle")
    worksheet.cell(row=1, column=2, value="Followers Count")
    worksheet.cell(row=1, column=3, value="Following Count")

    # Loop through each Twitter handle and get the follower and following counts
    for i, handle in enumerate(twitter_handles):
        user = api.get_user(screen_name=handle)
        row = i + 2
        worksheet.cell(row=row, column=1, value=user.screen_name)
        worksheet.cell(row=row, column=2, value=user.followers_count)
        worksheet.cell(row=row, column=3, value=user.friends_count)

    # Save the Excel workbook with the current date and time as the filename
    filename = datetime.datetime.now().strftime("twitter_followers_following_%Y-%m-%d_%H-%M-%S.xlsx")
    workbook.save(filename)

    # Send the Excel file as an attachment via email
    email_subject = "Twitter follower and following counts"
    email_body = "Please find attached the latest Twitter follower and following counts."
    attachment_path = filename

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = email_subject
    message.attach(MIMEText(email_body))

    with open(attachment_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename= {filename}")
        message.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as smtp_server:
        smtp_server.starttls()
        smtp_server.login(sender_email, sender_password)
        smtp_server.sendmail(sender_email, receiver_email, message.as_string())

# Schedule the script to run every minute
# Disable the comment to run every min during the testing to make sure everything works 
# Do Not forget to coment out the everyday Schedual code during testing

#schedule.every(1).minutes.do(get_follower_counts)

# Schedule the script to run every day at 11:30
schedule.every().day.at("23:30").do(get_follower_counts)
while True:
    schedule.run_pending()
    time.sleep(1)
